import streamlit as st
import anthropic
import base64
import json
import io
import re
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side
)
from openpyxl.utils import get_column_letter
 
 
# ── helpers ──────────────────────────────────────────────────────────────────
 
def pdf_to_base64(uploaded_file) -> str:
    return base64.standard_b64encode(uploaded_file.read()).decode("utf-8")
 
 
def image_to_base64(uploaded_file) -> str:
    return base64.standard_b64encode(uploaded_file.read()).decode("utf-8")
 
 
def extract_lvc_data(file_bytes: bytes, media_type: str) -> dict:
    """Send the PDF/image to Claude and get structured LVC data back."""
    client = anthropic.Anthropic()
 
    system_prompt = """Você é um especialista em leitura de fichas de Levantamento Visual de Campo (LVC) 
de rodovias brasileiras. Sua tarefa é extrair TODOS os dados da ficha com precisão absoluta.
 
Retorne SOMENTE um JSON válido, sem texto adicional, no seguinte formato:
{
  "data": "DD/MM/AAAA ou vazio",
  "rodovia": "código da rodovia",
  "equipe": "nomes da equipe",
  "trecho": "descrição do trecho",
  "inicio_ponto": "ponto inicial",
  "fim_ponto": "ponto final",
  "extensao_km": número ou null,
  "segmentos": [
    {
      "km_ini": número,
      "km_fim": número,
      "P": true/false,
      "A": true/false,
      "S": true/false,
      "E": true/false,
      "D": true/false,
      "observacoes": "texto ou vazio",
      "ponto": "código do ponto ou vazio",
      "foto": "referência de foto ou vazio"
    }
  ]
}
 
INSTRUÇÕES CRÍTICAS:
- Leia CADA linha da tabela, do km 0-1 até o último km registrado
- Um X (ou marca similar) na célula = true; célula vazia = false
- Preserve o texto exato das observações, pontos e fotos
- Se uma coluna não existir na ficha, use false para todos os segmentos
- Não invente dados; se não conseguir ler, use null ou string vazia"""
 
    if media_type == "application/pdf":
        content_block = {
            "type": "document",
            "source": {
                "type": "base64",
                "media_type": "application/pdf",
                "data": base64.standard_b64encode(file_bytes).decode("utf-8"),
            },
        }
    else:
        content_block = {
            "type": "image",
            "source": {
                "type": "base64",
                "media_type": media_type,
                "data": base64.standard_b64encode(file_bytes).decode("utf-8"),
            },
        }
 
    response = client.messages.create(
        model="claude-opus-4-5",
        max_tokens=4096,
        system=system_prompt,
        messages=[
            {
                "role": "user",
                "content": [
                    content_block,
                    {
                        "type": "text",
                        "text": "Extraia todos os dados desta ficha LVC e retorne o JSON conforme instruído.",
                    },
                ],
            }
        ],
    )
 
    raw = response.content[0].text.strip()
    # strip possible markdown fences
    raw = re.sub(r"^```(?:json)?\s*", "", raw)
    raw = re.sub(r"\s*```$", "", raw)
    return json.loads(raw)
 
 
# ── Excel builder ─────────────────────────────────────────────────────────────
 
def _border(style="thin"):
    s = Side(style=style)
    return Border(left=s, right=s, top=s, bottom=s)
 
 
def _fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)
 
 
def build_excel(data: dict) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "LVC"
 
    # ── palette
    HEADER_BG   = "1F4E79"
    HEADER_FG   = "FFFFFF"
    SUB_BG      = "D6E4F0"
    MARK_BG     = "FF0000"
    MARK_FG     = "FFFFFF"
    ALT_ROW_BG  = "EBF5FB"
    BORDER_COL  = "2E75B6"
 
    thin_border = _border("thin")
 
    def hdr_cell(cell, value, bg=HEADER_BG, fg=HEADER_FG, size=11, bold=True, wrap=False):
        cell.value = value
        cell.font = Font(name="Arial", bold=bold, color=fg, size=size)
        cell.fill = _fill(bg)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=wrap)
        cell.border = Border(
            left=Side(style="medium", color=BORDER_COL),
            right=Side(style="medium", color=BORDER_COL),
            top=Side(style="medium", color=BORDER_COL),
            bottom=Side(style="medium", color=BORDER_COL),
        )
 
    # ── title row
    ws.merge_cells("A1:J1")
    title_cell = ws["A1"]
    title_cell.value = "FICHA DE LEVANTAMENTO VISUAL - TRAFEGABILIDADE"
    title_cell.font = Font(name="Arial", bold=True, size=13, color=HEADER_FG)
    title_cell.fill = _fill(HEADER_BG)
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    title_cell.border = _border("medium")
    ws.row_dimensions[1].height = 24
 
    # ── metadata rows (rows 2-4)
    meta = [
        ("DATA:", data.get("data", ""), "EQUIPE:", data.get("equipe", "")),
        ("RODOVIA:", data.get("rodovia", ""), "TRECHO:", data.get("trecho", "")),
        ("INÍCIO / PONTO:", data.get("inicio_ponto", ""), "EXT. (km):", data.get("extensao_km", "")),
        ("FIM / PONTO:", data.get("fim_ponto", ""), "", ""),
    ]
 
    for r_offset, (lbl1, val1, lbl2, val2) in enumerate(meta, start=2):
        row = r_offset
        ws.row_dimensions[row].height = 18
        for col, (val, is_label) in enumerate(
            [(lbl1, True), (val1, False), (lbl2, True), (val2, False)], start=1
        ):
            if col == 1:
                cells = [ws.cell(row=row, column=1)]
                ws.merge_cells(f"A{row}:B{row}")
                c = ws.cell(row=row, column=1)
            elif col == 2:
                c = ws.cell(row=row, column=3)
                ws.merge_cells(f"C{row}:E{row}")
            elif col == 3:
                c = ws.cell(row=row, column=6)
                ws.merge_cells(f"F{row}:G{row}")
            else:
                c = ws.cell(row=row, column=8)
                ws.merge_cells(f"H{row}:J{row}")
 
            c.value = val
            c.font = Font(name="Arial", bold=is_label, size=10,
                          color=HEADER_FG if is_label else "000000")
            c.fill = _fill(HEADER_BG if is_label else "FFFFFF")
            c.alignment = Alignment(horizontal="left" if not is_label else "right",
                                    vertical="center")
            c.border = thin_border
 
    # ── column headers (row 6)
    hdr_row = 6
    ws.row_dimensions[hdr_row].height = 30
    headers = ["KM INI", "KM FIM", "(P)\nPANELA", "(A)\nAFUND.", "(S)\nSINAL.",
               "(E)\nEROSÃO", "(D)\nDESLIZ.", "OBSERVAÇÕES", "PONTO", "FOTO"]
    widths  = [8, 8, 9, 9, 9, 9, 9, 38, 10, 12]
    cols    = list("ABCDEFGHIJ")
 
    for i, (h, w, col) in enumerate(zip(headers, widths, cols), start=1):
        ws.column_dimensions[col].width = w
        c = ws.cell(row=hdr_row, column=i)
        hdr_cell(c, h, wrap=True)
 
    # ── data rows
    segmentos = data.get("segmentos", [])
    for s_idx, seg in enumerate(segmentos):
        row = hdr_row + 1 + s_idx
        ws.row_dimensions[row].height = 17
        alt = s_idx % 2 == 1
 
        base_bg = ALT_ROW_BG if alt else "FFFFFF"
 
        values = [
            seg.get("km_ini", ""),
            seg.get("km_fim", ""),
            seg.get("P", False),
            seg.get("A", False),
            seg.get("S", False),
            seg.get("E", False),
            seg.get("D", False),
            seg.get("observacoes", ""),
            seg.get("ponto", ""),
            seg.get("foto", ""),
        ]
 
        for col_i, val in enumerate(values, start=1):
            c = ws.cell(row=row, column=col_i)
            c.border = thin_border
 
            if isinstance(val, bool):
                if val:
                    c.value = "✗"
                    c.font = Font(name="Arial", bold=True, color=MARK_FG, size=12)
                    c.fill = _fill(MARK_BG)
                else:
                    c.value = ""
                    c.fill = _fill(base_bg)
                c.alignment = Alignment(horizontal="center", vertical="center")
            else:
                c.value = val
                c.fill = _fill(base_bg)
                c.font = Font(name="Arial", size=10)
                if col_i <= 2:
                    c.alignment = Alignment(horizontal="center", vertical="center")
                else:
                    c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
 
    # ── totals row
    last_data_row = hdr_row + len(segmentos)
    total_row = last_data_row + 2
    ws.row_dimensions[total_row].height = 18
    ws.merge_cells(f"A{total_row}:B{total_row}")
    tc = ws.cell(row=total_row, column=1)
    hdr_cell(tc, "TOTAIS", bg=SUB_BG.replace("D6E4F0", "2E75B6"), fg=HEADER_FG)
 
    for col_i, col_letter in enumerate(["C", "D", "E", "F", "G"], start=3):
        c = ws.cell(row=total_row, column=col_i)
        first = hdr_row + 1
        last  = last_data_row
        c.value = f'=COUNTIF({col_letter}{first}:{col_letter}{last},"✗")'
        c.font  = Font(name="Arial", bold=True, size=10, color=HEADER_FG)
        c.fill  = _fill("2E75B6")
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = _border("medium")
 
    # ── legend
    legend_row = total_row + 2
    ws.merge_cells(f"A{legend_row}:J{legend_row}")
    lc = ws.cell(row=legend_row, column=1)
    lc.value = ("LEGENDA:  (P) PANELA  |  (A) AFUNDAMENTO  |  (S) SINALIZAÇÃO DEFICIENTE  |"
                "  (E) EROSÃO  |  (D) DESLIZAMENTO")
    lc.font  = Font(name="Arial", italic=True, size=9, color="444444")
    lc.alignment = Alignment(horizontal="left", vertical="center")
 
    obs_row = legend_row + 1
    ws.merge_cells(f"A{obs_row}:J{obs_row}")
    oc = ws.cell(row=obs_row, column=1)
    oc.value = "OBSERVAÇÕES: REGISTRO DE OUTRAS SITUAÇÕES DE RISCO E/OU DE SEGMENTOS CRÍTICOS"
    oc.font  = Font(name="Arial", italic=True, size=9, color="444444")
    oc.alignment = Alignment(horizontal="left", vertical="center")
 
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()
 
 
# ── Streamlit page ────────────────────────────────────────────────────────────
 
def render():
    st.title("📋 Leitura de Fichas LVC")
    st.markdown(
        "Faça o upload de uma ou mais fichas de **Levantamento Visual de Campo** "
        "(PDF ou imagem) e baixe a planilha Excel formatada."
    )
 
    uploaded_files = st.file_uploader(
        "Selecione as fichas LVC",
        type=["pdf", "jpg", "jpeg", "png"],
        accept_multiple_files=True,
    )
 
    if not uploaded_files:
        st.info("⬆️ Faça o upload das fichas para começar.")
        return
 
    if st.button("🔍 Processar fichas", type="primary"):
        all_results = []
 
        for uploaded_file in uploaded_files:
            with st.spinner(f"Lendo **{uploaded_file.name}**…"):
                try:
                    file_bytes = uploaded_file.read()
                    ext = uploaded_file.name.rsplit(".", 1)[-1].lower()
                    media_type_map = {
                        "pdf": "application/pdf",
                        "jpg": "image/jpeg",
                        "jpeg": "image/jpeg",
                        "png": "image/png",
                    }
                    media_type = media_type_map.get(ext, "application/pdf")
                    data = extract_lvc_data(file_bytes, media_type)
                    all_results.append((uploaded_file.name, data))
                    st.success(f"✅ {uploaded_file.name} — {len(data.get('segmentos', []))} segmentos lidos")
                except Exception as e:
                    st.error(f"❌ Erro ao processar {uploaded_file.name}: {e}")
 
        if not all_results:
            return
 
        # preview + download per file
        for fname, data in all_results:
            with st.expander(f"📄 Resultado: {fname}", expanded=True):
                col1, col2 = st.columns(2)
                with col1:
                    st.markdown(f"**Rodovia:** {data.get('rodovia','—')}")
                    st.markdown(f"**Trecho:** {data.get('trecho','—')}")
                    st.markdown(f"**Equipe:** {data.get('equipe','—')}")
                with col2:
                    st.markdown(f"**Data:** {data.get('data','—')}")
                    st.markdown(f"**Extensão:** {data.get('extensao_km','—')} km")
                    segs = data.get("segmentos", [])
                    st.markdown(f"**Segmentos:** {len(segs)}")
 
                segs = data.get("segmentos", [])
                if segs:
                    import pandas as pd
                    df = pd.DataFrame(segs)
                    bool_cols = ["P", "A", "S", "E", "D"]
                    for col in bool_cols:
                        if col in df.columns:
                            df[col] = df[col].apply(lambda x: "✗" if x else "")
                    st.dataframe(df, use_container_width=True, hide_index=True)
 
                    xlsx_bytes = build_excel(data)
                    base_name = fname.rsplit(".", 1)[0]
                    st.download_button(
                        label="📥 Baixar Excel",
                        data=xlsx_bytes,
                        file_name=f"LVC_{base_name}_{datetime.today().strftime('%Y%m%d')}.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    )
 
        # optional: merged workbook when multiple files
        if len(all_results) > 1:
            st.divider()
            st.subheader("📦 Download consolidado (todas as fichas)")
            wb_all = Workbook()
            wb_all.remove(wb_all.active)
            for fname, data in all_results:
                sheet_name = fname.rsplit(".", 1)[0][:31]
                _add_sheet_to_workbook(wb_all, sheet_name, data)
 
            buf = io.BytesIO()
            wb_all.save(buf)
            buf.seek(0)
            st.download_button(
                label="📥 Baixar Excel consolidado",
                data=buf.getvalue(),
                file_name=f"LVC_consolidado_{datetime.today().strftime('%Y%m%d')}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
 
 
def _add_sheet_to_workbook(wb: Workbook, sheet_name: str, data: dict):
    """Clone the formatting logic into an existing workbook as a new sheet."""
    buf_bytes = build_excel(data)
    from openpyxl import load_workbook
    tmp_wb = load_workbook(io.BytesIO(buf_bytes))
    tmp_ws = tmp_wb.active
    new_ws = wb.create_sheet(title=sheet_name)
    for row in tmp_ws.iter_rows():
        for cell in row:
            nc = new_ws.cell(row=cell.row, column=cell.column, value=cell.value)
            if cell.has_style:
                nc.font      = cell.font.copy()
                nc.fill      = cell.fill.copy()
                nc.border    = cell.border.copy()
                nc.alignment = cell.alignment.copy()
    for col, dim in tmp_ws.column_dimensions.items():
        new_ws.column_dimensions[col].width = dim.width
    for row_num, dim in tmp_ws.row_dimensions.items():
        new_ws.row_dimensions[row_num].height = dim.height
    for merge in tmp_ws.merged_cells.ranges:
        new_ws.merge_cells(str(merge))
 
 
# ── entrypoint ────────────────────────────────────────────────────────────────
 
if __name__ == "__main__":
    render()
